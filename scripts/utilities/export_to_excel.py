import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

csv_path = 'features/found_pieces.csv'
excel_path = 'features/found_pieces.xlsx'

print(f"Reading {csv_path}...")
df = pd.read_csv(csv_path)

# Insert 'validated' column at the beginning (index 0)
df.insert(0, 'validated', '')

print(f"Saving to {excel_path}...")
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Found Pieces')
    
    # Access the workbook and sheet to style it
    workbook = writer.book
    worksheet = writer.sheets['Found Pieces']
    
    # Enable grid lines explicitly
    worksheet.views.sheetView[0].showGridLines = True
    
    # Header styling: bold, blue fill, white text, centered
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Classic steel/navy blue
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Style header row
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    # Style data rows
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = Font(name='Arial', size=10)
            cell.border = thin_border
            # Align 'validated' column centered
            if col == 1:
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Determine the maximum width
        for cell in col:
            val = str(cell.value or '')
            if cell.row == 1:
                max_len = max(max_len, len(val) + 4)
            else:
                max_len = max(max_len, len(val))
        
        # Set column width with padding
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

print("Done exporting and styling!")
